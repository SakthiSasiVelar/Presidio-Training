# Create a application that will take the Employee details(Name, DOB, Phone and E-Mail) from console, validate it and calculate age(Age should not taken from user)
# The application should show menu to store the same in file. Option for saving should be text/excel/pdf. 


import datetime
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from fpdf import FPDF

class Employee :
    def __init__(self, userDetails):
        self.userDetails = userDetails
    
    def __str__(self):
        return self.userDetails
    



def get_input_from_user():
    name = input('Enter Name: ')
    dob = input('Enter Date of Birth(YYYY-MM-DD): ')
    phone = input('Enter Phone Number: ')
    email = input('Enter Email: ')
   
    userDetails = {
        'name': name,
        'dob': dob,
        'phone': phone,
        'email': email
    }
    return userDetails

def validate_user_details(userDetails):
    if not userDetails['name'].isalpha():
        print('Invalid Name. Name should only contain alphabets.')
        return False
    try:
        dob = datetime.datetime.strptime(userDetails['dob'], '%Y-%m-%d')
        if dob > datetime.datetime.now():
            print('Invalid Date of Birth. Date should be in the past.')
            return False
    except ValueError:
        print('Invalid Date of Birth. Please enter date in YYYY-MM-DD format.')
        return False
    
    if not userDetails['phone'].isdigit() or len(userDetails['phone'])!= 10:
        print('Invalid Phone Number. Phone number should be 10 digits.')
        return False
    
    if '@' not in userDetails['email'] or '.' not in userDetails['email']:
        print('Invalid Email. Email should contain "@" and "."')
        return False
    
    return True

def calculate_age(dob):
    birthDate = datetime.datetime.strptime(dob, "%Y-%m-%d")
    today = datetime.date.today()
    age = today.year - birthDate.year - ((today.month, today.day) < (birthDate.month, birthDate.day))
    return age

def save_to_text_file(userDetails):
       with open("employees.txt", "w") as file:
            for key, value in userDetails.items():
                file.write(f"{key}: {value}\n")

def save_to_excel_file(userDetails):
    df = pd.DataFrame(userDetails , index=[0])
    df.to_excel('employees.xlsx', index=False)

def save_to_pdf_file(userDetails):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Employee Details', 0, 1, 'C')
    pdf.set_font('Arial', '', 12)
    for key, value in userDetails.items():
        pdf.cell(0, 10, f'{key}: {value}', 0, 1)
    pdf.output('employees.pdf')

def bulk_read_from_excel(file_path):
    employees = []
    wb = load_workbook(filename=file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        employee = {
            'name': row[0],
            'email': row[1],
            'age': int(row[2]),
        }
        employees.append(employee)
    print (employees)

def program():
    userDetails = get_input_from_user()
    if(validate_user_details(userDetails)) :
        age = calculate_age(userDetails['dob'])
        userDetails['age'] = age
        employee = Employee(userDetails)
        print("Choose an option to save the details:")
        print("1. Save to Text File")
        print("2. Save to Excel File")
        print("3. Save to PDF File")
        print("4. Bulk Read from Excel")
        choice = int(input("Enter your choice (1/2/3/4): "))

        if choice == 1:
            save_to_text_file (employee.userDetails)
            print("Details saved successfully to Text File.")
        elif choice == 2:
            save_to_excel_file (employee.userDetails)
            print("Details saved successfully to Excel File.")
        elif choice == 3:
            save_to_pdf_file(employee.userDetails)
            print("Details saved successfully to PDF File.")
        elif choice == 4:
            bulk_read_from_excel('employeeDetails.xlsx')
            print("Bulk read from Excel successfully.")
            
program()